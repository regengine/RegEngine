from __future__ import annotations

import csv
import io

from fastapi import HTTPException


CSV_HEADERS = [
    "record_type",
    "facility_name",
    "name",
    "street",
    "city",
    "state",
    "postal_code",
    "fda_registration_number",
    "roles",
    "category_id",
    "tlc_code",
    "product_description",
    "status",
    "cte_type",
    "event_time",
    "kde_data",
    "obligation_ids",
]


def _csv_template_bytes() -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(CSV_HEADERS)
    writer.writerow(["facility", "", "Salinas Packhouse", "1200 Abbott St", "Salinas", "CA", "93901", "12345678901", "Grower,Packer", "", "", "", "", "", "", "", ""])
    writer.writerow(["ftl_scope", "Salinas Packhouse", "", "", "", "", "", "", "", "2", "", "", "", "", "", "", ""])
    writer.writerow(["tlc", "Salinas Packhouse", "", "", "", "", "", "", "", "", "TLC-2026-SAL-1001", "Baby Spinach", "active", "", "", "", ""])
    writer.writerow(["event", "Salinas Packhouse", "", "", "", "", "", "", "", "", "TLC-2026-SAL-1001", "", "", "shipping", "2026-03-03T12:00:00Z", '{"quantity": 120, "unit_of_measure": "cases", "reference_document": "BOL-1001"}', ""])
    return buffer.getvalue().encode("utf-8")


def _xlsx_template_bytes() -> bytes:
    try:
        from openpyxl import Workbook
    except Exception as exc:  # pragma: no cover - dependency/environment specific
        raise HTTPException(status_code=503, detail="XLSX template generation unavailable (openpyxl missing)") from exc

    workbook = Workbook()
    facilities_sheet = workbook.active
    facilities_sheet.title = "facilities"
    facilities_sheet.append(["name", "street", "city", "state", "postal_code", "fda_registration_number", "roles"])
    facilities_sheet.append(["Salinas Packhouse", "1200 Abbott St", "Salinas", "CA", "93901", "12345678901", "Grower,Packer"])

    ftl_sheet = workbook.create_sheet("ftl_scope")
    ftl_sheet.append(["facility_name", "category_id"])
    ftl_sheet.append(["Salinas Packhouse", "2"])

    tlc_sheet = workbook.create_sheet("tlcs")
    tlc_sheet.append(["facility_name", "tlc_code", "product_description", "status"])
    tlc_sheet.append(["Salinas Packhouse", "TLC-2026-SAL-1001", "Baby Spinach", "active"])

    cte_sheet = workbook.create_sheet("cte_events")
    cte_sheet.append(["facility_name", "tlc_code", "cte_type", "event_time", "kde_data", "obligation_ids"])
    cte_sheet.append(
        [
            "Salinas Packhouse",
            "TLC-2026-SAL-1001",
            "shipping",
            "2026-03-03T12:00:00Z",
            '{"quantity": 120, "unit_of_measure": "cases", "reference_document": "BOL-1001"}',
            "",
        ]
    )

    reference_sheet = workbook.create_sheet("reference")
    reference_sheet.append(["FTL category_id", "Description"]) 
    for category_id, description in [
        ("1", "Fruits (fresh-cut)"),
        ("2", "Vegetables (leafy greens)"),
        ("3", "Shell eggs"),
        ("4", "Nut butter"),
        ("5", "Fresh herbs"),
        ("6", "Finfish (fresh/frozen)"),
        ("7", "Crustaceans (fresh/frozen)"),
        ("8", "Molluscan shellfish"),
        ("9", "Ready-to-eat deli salads"),
        ("10", "Soft & semi-soft cheeses"),
    ]:
        reference_sheet.append([category_id, description])

    reference_sheet.append([])
    reference_sheet.append(["Supported cte_type values", ""]) 
    for cte_type in [
        "shipping",
        "receiving",
        "transforming",
        "harvesting",
        "cooling",
        "initial_packing",
        "first_receiver",
    ]:
        reference_sheet.append([cte_type, ""]) 

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_template(format_name: str) -> tuple[bytes, str, str]:
    normalized = (format_name or "csv").strip().lower()
    if normalized == "csv":
        return _csv_template_bytes(), "text/csv", "supplier_bulk_upload_template.csv"
    if normalized == "xlsx":
        return (
            _xlsx_template_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "supplier_bulk_upload_template.xlsx",
        )
    raise HTTPException(status_code=400, detail="format must be csv or xlsx")
