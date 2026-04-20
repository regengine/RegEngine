from __future__ import annotations

import csv
import io
import json
import logging
import os
import re
from typing import Any

from fastapi import HTTPException, UploadFile

logger = logging.getLogger("bulk_upload.parsers")


ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".json", ".pdf"}
DEFAULT_MAX_UPLOAD_BYTES = 10 * 1024 * 1024
READ_CHUNK_BYTES = 1024 * 1024
DEFAULT_MAX_PDF_PAGES = 200


def _max_upload_bytes() -> int:
    raw_value = os.getenv(
        "SUPPLIER_BULK_UPLOAD_MAX_BYTES", str(DEFAULT_MAX_UPLOAD_BYTES)
    ).strip()
    try:
        parsed = int(raw_value)
    except ValueError:
        return DEFAULT_MAX_UPLOAD_BYTES
    if parsed <= 0:
        return DEFAULT_MAX_UPLOAD_BYTES
    return parsed


def _max_pdf_pages() -> int:
    """Hard cap on pages in an uploaded PDF to prevent DoS via PDF bombs.

    A compressed PDF can decompress into tens of thousands of pages; each
    ``extract_tables()`` / ``extract_text()`` call on such a document pins
    a worker at ~100% CPU and inflates RSS for minutes. The upload-bytes
    cap bounds wire size only — post-decompression page count must also
    be bounded before iteration begins.
    """
    raw_value = os.getenv(
        "BULK_UPLOAD_MAX_PDF_PAGES", str(DEFAULT_MAX_PDF_PAGES)
    ).strip()
    try:
        parsed = int(raw_value)
    except ValueError:
        return DEFAULT_MAX_PDF_PAGES
    if parsed <= 0:
        return DEFAULT_MAX_PDF_PAGES
    return parsed


WIDE_TABLE_ATTRIBUTE_MARKERS = {
    "storage_capacity",
    "pallet_positions",
    "temp_range",
    "loading_dock",
    "special_controls",
    "bonded_facility",
    "usda_inspection",
    "ppq_inspection",
    "overweight",
    "rail_access",
    "distance_from_port",
    "available_office_space",
    "contact",
    "additional_service",
}

PHONE_RE = re.compile(r"\(?\d{3}\)?[\s.-]\d{3}[\s.-]\d{4}")
STATE_ZIP_RE = re.compile(r"\b([A-Z]{2})\s*(\d{5}(?:-\d{4})?)\b")
FULL_ADDRESS_RE = re.compile(
    r"(?P<street>\d[^,\n]+?)\s*,\s*(?P<city>[A-Za-z .'-]+?)\s*,\s*(?P<state>[A-Z]{2})\s*(?P<postal>\d{5}(?:-\d{4})?)"
)


def _normalize_key(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", value.strip().lower())
    return normalized.strip("_")


# Canonical column alias map — all known variants collapse to internal names
# at ingestion so downstream code only references canonical keys.
COLUMN_ALIASES: dict[str, str] = {
    # facility_name
    "location_name": "facility_name",
    "origin_facility_name": "facility_name",
    "destination_facility_name": "facility_name",
    "business_name": "facility_name",
    "site_name": "facility_name",
    "sender_name": "facility_name",
    "receiver_name": "facility_name",
    "source": "facility_name",
    "destination": "facility_name",
    # tlc_code
    "traceability_lot_code": "tlc_code",
    "lot_number": "tlc_code",
    "lot_code": "tlc_code",
    "tlc": "tlc_code",
    # cte_type
    "event_type": "cte_type",
    # event_time
    "event_datetime": "event_time",
    "event_date": "event_time",
    "cte_date": "event_time",
    "timestamp": "event_time",
}


def _clean_multiline_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    lines = []
    for raw_line in text.split("\n"):
        line = re.sub(r"\s+", " ", raw_line).strip(" ,;\t")
        if line:
            lines.append(line)
    return "\n".join(lines)


def _extract_address_parts(text: str) -> tuple[str, str, str, str]:
    blob = _clean_multiline_text(text)
    if not blob:
        return "", "", "", ""

    one_line = blob.replace("\n", " ")
    address_match = FULL_ADDRESS_RE.search(one_line)
    if address_match:
        return (
            address_match.group("street").strip(),
            address_match.group("city").strip(),
            address_match.group("state").strip().upper(),
            address_match.group("postal").strip(),
        )

    street = ""
    city = ""
    state = ""
    postal_code = ""

    lines = [line for line in blob.split("\n") if line]
    for index, line in enumerate(lines):
        upper_line = line.upper()
        state_zip_match = STATE_ZIP_RE.search(upper_line)
        if state_zip_match:
            state = state_zip_match.group(1)
            postal_code = state_zip_match.group(2)
            city_part = line[: state_zip_match.start()].strip(" ,")
            if "," in city_part:
                city = city_part.split(",")[-1].strip()
            else:
                city = city_part.strip()
            if not street and index > 0:
                previous_line = lines[index - 1]
                if not PHONE_RE.search(previous_line):
                    street = previous_line
            break

    if not street:
        for line in lines:
            if PHONE_RE.search(line):
                continue
            lowered = line.lower()
            if lowered.startswith("www") or lowered.startswith("http"):
                continue
            if re.search(r"\d", line):
                street = line
                break

    return street, city, state, postal_code


def _infer_roles_from_text(text: str) -> list[str]:
    lowered = text.lower()
    roles: list[str] = []
    if any(token in lowered for token in ["grow", "harvest", "farm"]):
        roles.append("Grower")
    if any(token in lowered for token in ["pack", "re-pack", "bagging"]):
        roles.append("Packer")
    if any(
        token in lowered for token in ["process", "blast freez", "cooling", "transform"]
    ):
        roles.append("Processor")
    if any(
        token in lowered
        for token in [
            "drayage",
            "warehouse",
            "distribution",
            "shipping",
            "cross docking",
            "trans-loading",
        ]
    ):
        roles.append("Distributor")
    if any(token in lowered for token in ["import", "customs", "broker"]):
        roles.append("Importer")
    return sorted(set(roles))


def _extract_matrix_facility(
    header: Any,
    attributes: dict[str, str],
    *,
    warnings: list[str],
) -> dict[str, Any] | None:
    name = _clean_multiline_text(header).replace("\n", " ").strip()
    if not name:
        return None

    contact_blob = attributes.get("contact", "")
    additional_blob = attributes.get("additional_service", "")
    special_controls_blob = attributes.get("special_controls", "")
    lookup_blob = "\n".join(
        [name, contact_blob, additional_blob, special_controls_blob]
    )

    street, city, state, postal_code = _extract_address_parts(contact_blob)
    if not (street and city and state and postal_code):
        street_fallback, city_fallback, state_fallback, postal_fallback = (
            _extract_address_parts(lookup_blob)
        )
        street = street or street_fallback
        city = city or city_fallback
        state = state or state_fallback
        postal_code = postal_code or postal_fallback

    if not (street and city and state and postal_code):
        warnings.append(
            f"Skipped facility '{name}' because address fields were not detected"
        )
        return None

    return {
        "name": name,
        "street": street,
        "city": city,
        "state": state,
        "postal_code": postal_code,
        "fda_registration_number": None,
        "roles": _infer_roles_from_text(lookup_blob),
    }


def _parse_pdf_matrix_table(
    table: list[list[Any]], parsed: dict[str, Any], warnings: list[str]
) -> int:
    if not table or len(table) < 3:
        return 0

    row_labels: list[str] = []
    for row in table[1:]:
        if not row:
            continue
        label = _normalize_key(str(row[0] or ""))
        if label:
            row_labels.append(label)

    marker_hits = sum(
        1 for label in row_labels if label in WIDE_TABLE_ATTRIBUTE_MARKERS
    )
    if marker_hits < 4:
        return 0

    added = 0
    seen_names: set[str] = {
        str(item.get("name") or "").strip().lower()
        for item in parsed.get("facilities") or []
    }
    for column_index, header in enumerate(table[0][1:], start=1):
        attributes: dict[str, str] = {}
        for row in table[1:]:
            if not row:
                continue
            label = _normalize_key(str(row[0] or ""))
            if not label:
                continue
            if column_index >= len(row):
                continue
            value = _clean_multiline_text(row[column_index])
            if value:
                attributes[label] = value

        facility = _extract_matrix_facility(header, attributes, warnings=warnings)
        if facility is None:
            continue

        dedupe_key = str(facility.get("name") or "").strip().lower()
        if not dedupe_key or dedupe_key in seen_names:
            continue
        parsed["facilities"].append(facility)
        seen_names.add(dedupe_key)
        added += 1

    return added


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in row.items():
        if key is None:
            continue
        canonical = _normalize_key(str(key))
        canonical = COLUMN_ALIASES.get(canonical, canonical)
        # First writer wins — don't overwrite a populated canonical key
        # with a later alias that may be empty.
        if canonical in normalized and normalized[canonical]:
            continue
        normalized[canonical] = value
    return normalized


def _parse_json_field(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return {}
        try:
            parsed = json.loads(text)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            return {}
    return {}


def _parse_list_field(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        if text.startswith("["):
            try:
                parsed = json.loads(text)
                if isinstance(parsed, list):
                    return [str(item).strip() for item in parsed if str(item).strip()]
            except json.JSONDecodeError:
                pass
        return [token.strip() for token in text.split(",") if token.strip()]
    return []


def _extract_event_row(row: dict[str, Any]) -> dict[str, Any]:
    cte_type = row.get("cte_type") or ""
    event_time = row.get("event_time")
    kde_data = _parse_json_field(row.get("kde_data"))

    passthrough_keys = {
        "record_type",
        "facility_name",
        "tlc_code",
        "cte_type",
        "event_type",
        "event_time",
        "timestamp",
        "kde_data",
        "obligation_ids",
    }
    for key, value in row.items():
        if key in passthrough_keys:
            continue
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        kde_data[key] = value

    return {
        "facility_name": str(row.get("facility_name") or row.get("name") or "").strip(),
        "tlc_code": str(row.get("tlc_code") or "").strip(),
        "cte_type": str(cte_type).strip(),
        "event_time": str(event_time).strip() if event_time else None,
        "kde_data": kde_data,
        "obligation_ids": _parse_list_field(row.get("obligation_ids")),
    }


def _classify_row(
    parsed: dict[str, Any], row: dict[str, Any], *, warnings: list[str]
) -> None:
    record_type = str(row.get("record_type") or "").strip().lower()
    if record_type in {"facility", "facilities"}:
        parsed["facilities"].append(
            {
                "name": str(row.get("name") or row.get("facility_name") or "").strip(),
                "street": str(row.get("street") or "").strip(),
                "city": str(row.get("city") or "").strip(),
                "state": str(row.get("state") or "").strip(),
                "postal_code": str(
                    row.get("postal_code") or row.get("zip") or ""
                ).strip(),
                "fda_registration_number": str(
                    row.get("fda_registration_number") or ""
                ).strip()
                or None,
                "roles": _parse_list_field(row.get("roles")),
            }
        )
        return
    if record_type in {"ftl", "ftl_scope", "scope"}:
        parsed["ftl_scopes"].append(
            {
                "facility_name": str(
                    row.get("facility_name") or row.get("name") or ""
                ).strip(),
                "category_id": str(row.get("category_id") or "").strip(),
            }
        )
        return
    if record_type in {"tlc", "lot"}:
        parsed["tlcs"].append(
            {
                "tlc_code": str(row.get("tlc_code") or "").strip(),
                "facility_name": str(
                    row.get("facility_name") or row.get("name") or ""
                ).strip(),
                "product_description": str(row.get("product_description") or "").strip()
                or None,
                "status": str(row.get("status") or "active").strip().lower(),
            }
        )
        return
    if record_type in {"event", "cte", "cte_event"}:
        parsed["events"].append(_extract_event_row(row))
        return

    if row.get("cte_type"):
        parsed["events"].append(_extract_event_row(row))
        return
    if row.get("tlc_code"):
        parsed["events"].append(_extract_event_row(row))
        return
    if row.get("category_id"):
        parsed["ftl_scopes"].append(
            {
                "facility_name": str(
                    row.get("facility_name") or row.get("name") or ""
                ).strip(),
                "category_id": str(row.get("category_id") or "").strip(),
            }
        )
        return
    if row.get("tlc_code"):
        parsed["tlcs"].append(
            {
                "tlc_code": str(row.get("tlc_code") or "").strip(),
                "facility_name": str(
                    row.get("facility_name") or row.get("name") or ""
                ).strip(),
                "product_description": str(row.get("product_description") or "").strip()
                or None,
                "status": str(row.get("status") or "active").strip().lower(),
            }
        )
        return
    if row.get("name") or row.get("facility_name"):
        parsed["facilities"].append(
            {
                "name": str(row.get("name") or row.get("facility_name") or "").strip(),
                "street": str(row.get("street") or "").strip(),
                "city": str(row.get("city") or "").strip(),
                "state": str(row.get("state") or "").strip(),
                "postal_code": str(
                    row.get("postal_code") or row.get("zip") or ""
                ).strip(),
                "fda_registration_number": str(
                    row.get("fda_registration_number") or ""
                ).strip()
                or None,
                "roles": _parse_list_field(row.get("roles")),
            }
        )
        return

    warnings.append("Skipped unrecognized row in uploaded dataset")


def _parse_csv_bytes(
    content: bytes, parsed: dict[str, Any], warnings: list[str]
) -> None:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        normalized_row = _normalize_row(row)
        if not any(str(value or "").strip() for value in normalized_row.values()):
            continue
        _classify_row(parsed, normalized_row, warnings=warnings)


def _append_non_object_warning(
    warnings: list[str], section_name: str, index: int
) -> None:
    warnings.append(
        f"Skipped non-object row in {section_name} section at index {index}"
    )


def _parse_json_bytes(
    content: bytes, parsed: dict[str, Any], warnings: list[str]
) -> None:
    try:
        payload = json.loads(content.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise HTTPException(
            status_code=400, detail=f"Invalid JSON file: {exc}"
        ) from exc

    if isinstance(payload, dict):
        facilities = payload.get("facilities") or []
        ftl_scopes = payload.get("ftl_scopes") or payload.get("ftl_scope") or []
        tlcs = payload.get("tlcs") or []
        events = payload.get("events") or []

        if isinstance(facilities, list):
            for index, row in enumerate(facilities, start=1):
                if not isinstance(row, dict):
                    _append_non_object_warning(warnings, "facilities", index)
                    continue
                _classify_row(
                    parsed,
                    _normalize_row({**row, "record_type": "facility"}),
                    warnings=warnings,
                )
        if isinstance(ftl_scopes, list):
            for index, row in enumerate(ftl_scopes, start=1):
                if not isinstance(row, dict):
                    _append_non_object_warning(warnings, "ftl_scopes", index)
                    continue
                _classify_row(
                    parsed,
                    _normalize_row({**row, "record_type": "ftl_scope"}),
                    warnings=warnings,
                )
        if isinstance(tlcs, list):
            for index, row in enumerate(tlcs, start=1):
                if not isinstance(row, dict):
                    _append_non_object_warning(warnings, "tlcs", index)
                    continue
                _classify_row(
                    parsed,
                    _normalize_row({**row, "record_type": "tlc"}),
                    warnings=warnings,
                )
        if isinstance(events, list):
            for index, row in enumerate(events, start=1):
                if not isinstance(row, dict):
                    _append_non_object_warning(warnings, "events", index)
                    continue
                _classify_row(
                    parsed,
                    _normalize_row({**row, "record_type": "event"}),
                    warnings=warnings,
                )
        if not any(
            isinstance(section, list) and section
            for section in [facilities, ftl_scopes, tlcs, events]
        ):
            _classify_row(parsed, _normalize_row(payload), warnings=warnings)
        return

    if isinstance(payload, list):
        for index, row in enumerate(payload, start=1):
            if isinstance(row, dict):
                _classify_row(parsed, _normalize_row(row), warnings=warnings)
            else:
                _append_non_object_warning(warnings, "json_array", index)
        return

    raise HTTPException(status_code=400, detail="JSON must be an object or array")


def _parse_xlsx_bytes(
    content: bytes, parsed: dict[str, Any], warnings: list[str]
) -> None:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - dependency/env specific
        raise HTTPException(
            status_code=503, detail="XLSX parsing unavailable (openpyxl missing)"
        ) from exc

    workbook = load_workbook(io.BytesIO(content), data_only=True)
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [_normalize_key(str(cell or "")) for cell in rows[0]]
        if not any(headers):
            continue

        normalized_sheet_name = sheet.title.strip().lower()
        inferred_record_type: str | None = None
        if "facility" in normalized_sheet_name:
            inferred_record_type = "facility"
        elif "ftl" in normalized_sheet_name or "scope" in normalized_sheet_name:
            inferred_record_type = "ftl_scope"
        elif "tlc" in normalized_sheet_name or "lot" in normalized_sheet_name:
            inferred_record_type = "tlc"
        elif "event" in normalized_sheet_name or "cte" in normalized_sheet_name:
            inferred_record_type = "event"

        for values in rows[1:]:
            if values is None:
                continue
            row_data = {
                header: value for header, value in zip(headers, values) if header
            }
            if not any(str(value or "").strip() for value in row_data.values()):
                continue
            if inferred_record_type and "record_type" not in row_data:
                row_data["record_type"] = inferred_record_type
            _classify_row(parsed, row_data, warnings=warnings)


def _parse_pdf_bytes(
    content: bytes, parsed: dict[str, Any], warnings: list[str]
) -> None:
    try:
        import pdfplumber
    except Exception as exc:  # pragma: no cover - dependency/env specific
        raise HTTPException(
            status_code=503, detail="PDF parsing unavailable (pdfplumber missing)"
        ) from exc

    extracted_rows = 0
    max_pages = _max_pdf_pages()
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        page_count = len(pdf.pages)
        if page_count > max_pages:
            raise HTTPException(
                status_code=413,
                detail=(
                    f"PDF has {page_count} pages; max allowed is {max_pages}"
                ),
            )
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                if not table or len(table) < 2:
                    continue

                matrix_rows = _parse_pdf_matrix_table(table, parsed, warnings)
                if matrix_rows > 0:
                    extracted_rows += matrix_rows
                    continue

                headers = [_normalize_key(str(cell or "")) for cell in table[0]]
                for row_values in table[1:]:
                    row = {
                        header: value
                        for header, value in zip(headers, row_values)
                        if header
                    }
                    if not any(str(value or "").strip() for value in row.values()):
                        continue
                    _classify_row(parsed, _normalize_row(row), warnings=warnings)
                    extracted_rows += 1

        if extracted_rows == 0:
            full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            start = full_text.find("{")
            end = full_text.rfind("}")
            if 0 <= start < end:
                blob = full_text[start : end + 1]
                try:
                    _parse_json_bytes(blob.encode("utf-8"), parsed, warnings)
                    extracted_rows = 1
                except Exception:
                    logger.debug("PDF JSON-like text parsing failed", exc_info=True)
                    warnings.append("PDF contained JSON-like text but parsing failed")

    if extracted_rows == 0:
        warnings.append(
            "PDF parsed but no structured rows detected; include tabular data or JSON block"
        )


async def parse_incoming_file(file: UploadFile) -> dict[str, Any]:
    extension = os.path.splitext(file.filename or "")[1].lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Unsupported file format")

    max_bytes = _max_upload_bytes()
    chunks: list[bytes] = []
    total_bytes = 0

    while True:
        chunk = await file.read(READ_CHUNK_BYTES)
        if not chunk:
            break
        total_bytes += len(chunk)
        if total_bytes > max_bytes:
            raise HTTPException(
                status_code=413,
                detail=f"Uploaded file exceeds max size of {max_bytes} bytes",
            )
        chunks.append(chunk)

    content = b"".join(chunks)
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    parsed: dict[str, Any] = {
        "detected_format": extension.replace(".", ""),
        "facilities": [],
        "ftl_scopes": [],
        "tlcs": [],
        "events": [],
        "warnings": [],
    }
    warnings = parsed["warnings"]

    if extension == ".csv":
        _parse_csv_bytes(content, parsed, warnings)
    elif extension == ".json":
        _parse_json_bytes(content, parsed, warnings)
    elif extension == ".xlsx":
        _parse_xlsx_bytes(content, parsed, warnings)
    elif extension == ".pdf":
        _parse_pdf_bytes(content, parsed, warnings)
    else:  # pragma: no cover
        raise HTTPException(status_code=400, detail="Unsupported file format")

    _auto_scaffold_facilities(parsed, warnings)

    return parsed


def _auto_scaffold_facilities(
    parsed: dict[str, Any], warnings: list[str]
) -> None:
    """Auto-create facility records from event/tlc/ftl_scope rows.

    When a CSV contains event rows referencing facilities that have no
    explicit facility record, this function synthesises minimal facility
    entries from the data available in the event rows (city, state, etc.
    are captured in kde_data by the column-alias normaliser).
    """
    existing_names: set[str] = set()
    for facility in parsed.get("facilities") or []:
        name = str(facility.get("name") or "").strip().lower()
        if name:
            existing_names.add(name)

    # Collect unique facility names referenced by other sections and the
    # best available metadata for each.
    facility_meta: dict[str, dict[str, str]] = {}

    for section in ("events", "tlcs", "ftl_scopes"):
        for row in parsed.get(section) or []:
            raw_name = str(row.get("facility_name") or "").strip()
            if not raw_name:
                continue
            key = raw_name.lower()
            if key in existing_names:
                continue
            if key in facility_meta:
                continue

            # Pull city/state from kde_data where the parser stashed
            # non-passthrough columns like origin_city, origin_state.
            kde = row.get("kde_data") or {}
            city = str(
                kde.get("origin_city")
                or kde.get("city")
                or kde.get("destination_city")
                or kde.get("location_city")
                or ""
            ).strip()
            state = str(
                kde.get("origin_state")
                or kde.get("state")
                or kde.get("destination_state")
                or kde.get("location_state")
                or ""
            ).strip()

            facility_meta[key] = {
                "name": raw_name,
                "city": city,
                "state": state,
            }

    if not facility_meta:
        return

    for meta in facility_meta.values():
        parsed["facilities"].append(
            {
                "name": meta["name"],
                "street": "",
                "city": meta["city"],
                "state": meta["state"],
                "postal_code": "",
                "fda_registration_number": "",
                "roles": [],
            }
        )

    warnings.append(
        f"Auto-created {len(facility_meta)} facility record(s) from event data: "
        + ", ".join(m["name"] for m in facility_meta.values())
    )
