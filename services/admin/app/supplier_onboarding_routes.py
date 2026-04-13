from __future__ import annotations

import csv
import io
import uuid as uuid_module
from datetime import datetime, timedelta, timezone
from typing import Any
import zipfile

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from .database import get_session
from .dependencies import get_current_user
from .models import TenantContext
from .sqlalchemy_models import (
    SupplierCTEEventModel,
    SupplierFacilityFTLCategoryModel,
    SupplierFacilityModel,
    SupplierFunnelEventModel,
    SupplierTraceabilityLotModel,
    UserModel,
)
from .supplier_graph_sync import supplier_graph_sync
from .supplier_cte_service import (
    _persist_supplier_cte_event,
)


router = APIRouter(prefix="/supplier", tags=["supplier-onboarding"])


FTL_CATEGORY_CATALOG = [
    {"id": "1", "name": "Fruits (fresh-cut)", "ctes": ["receiving", "transforming", "shipping"]},
    {
        "id": "2",
        "name": "Vegetables (leafy greens)",
        "ctes": ["harvesting", "cooling", "initial_packing", "receiving", "transforming", "shipping"],
    },
    {"id": "3", "name": "Shell eggs", "ctes": ["initial_packing", "receiving", "shipping"]},
    {"id": "4", "name": "Nut butter", "ctes": ["receiving", "transforming", "shipping"]},
    {"id": "5", "name": "Fresh herbs", "ctes": ["harvesting", "cooling", "initial_packing", "receiving", "shipping"]},
    {"id": "6", "name": "Finfish (fresh/frozen)", "ctes": ["first_receiver", "receiving", "transforming", "shipping"]},
    {"id": "7", "name": "Crustaceans (fresh/frozen)", "ctes": ["first_receiver", "receiving", "transforming", "shipping"]},
    {"id": "8", "name": "Molluscan shellfish", "ctes": ["harvesting", "first_receiver", "receiving", "shipping"]},
    {"id": "9", "name": "Ready-to-eat deli salads", "ctes": ["receiving", "transforming", "shipping"]},
    {"id": "10", "name": "Soft & semi-soft cheeses", "ctes": ["receiving", "transforming", "shipping"]},
]

FTL_CATEGORY_LOOKUP = {item["id"]: item for item in FTL_CATEGORY_CATALOG}


class SupplierFacilityCreateRequest(BaseModel):
    name: str = Field(min_length=2)
    street: str = Field(min_length=2)
    city: str = Field(min_length=2)
    state: str = Field(min_length=2)
    postal_code: str = Field(min_length=3)
    fda_registration_number: str | None = None
    roles: list[str] = Field(default_factory=list)


class SupplierFacilityResponse(BaseModel):
    id: str
    name: str
    street: str
    city: str
    state: str
    postal_code: str
    fda_registration_number: str | None
    roles: list[str]


class FacilityFTLScopingRequest(BaseModel):
    category_ids: list[str] = Field(default_factory=list)


class FacilityFTLScopingResponse(BaseModel):
    facility_id: str
    categories: list[dict[str, Any]]
    required_ctes: list[str]
    source: str


class SupplierCTEEventCreateRequest(BaseModel):
    cte_type: str = Field(min_length=2)
    tlc_code: str = Field(min_length=3)
    event_time: datetime | None = None
    kde_data: dict[str, Any] = Field(default_factory=dict)
    obligation_ids: list[str] = Field(default_factory=list)


class SupplierCTEEventResponse(BaseModel):
    event_id: str
    facility_id: str
    tlc_code: str
    cte_type: str
    payload_sha256: str
    merkle_hash: str
    merkle_prev_hash: str | None
    merkle_sequence: int


class SupplierTLCUpsertRequest(BaseModel):
    facility_id: str
    tlc_code: str = Field(min_length=3)
    product_description: str | None = None
    status: str | None = "active"


class SupplierTLCResponse(BaseModel):
    id: str
    facility_id: str
    tlc_code: str
    product_description: str | None
    status: str
    event_count: int
    created_at: str


class SupplierComplianceScoreResponse(BaseModel):
    score: int
    coverage_ratio: float
    freshness_ratio: float
    integrity_ratio: float
    required_ctes: int
    covered_ctes: int
    stale_ctes: int
    missing_ctes: int
    total_events: int
    evaluated_at: str


class SupplierComplianceGap(BaseModel):
    facility_id: str
    facility_name: str
    cte_type: str | None
    severity: str
    issue: str
    reason: str
    last_seen: str | None = None


class SupplierComplianceGapsResponse(BaseModel):
    gaps: list[SupplierComplianceGap]
    total: int
    high: int
    medium: int
    low: int
    evaluated_at: str


class SupplierDemoResetResponse(BaseModel):
    focus_facility_id: str
    focus_facility_name: str
    focus_required_ctes: list[str]
    focus_gap_cte: str | None = None
    focus_gap_issue: str | None = None
    focus_gap_reason: str | None = None
    seeded_facilities: int
    seeded_tlcs: int
    seeded_tlc_codes: list[str]
    seeded_events: int
    seeded_cte_types: list[str]
    dashboard_score: int
    open_gap_count: int


class SupplierFunnelEventRequest(BaseModel):
    event_name: str = Field(min_length=3, max_length=80)
    step: str | None = None
    status: str | None = None
    facility_id: str | None = None
    metadata: dict[str, Any] = Field(default_factory=dict)


class SupplierFunnelEventResponse(BaseModel):
    event_id: str
    event_name: str
    created_at: str


class SupplierSocialProofResponse(BaseModel):
    suppliers_onboarded: int
    facilities_registered: int
    tlcs_tracked: int
    cte_events_verified: int
    fda_exports_generated: int
    updated_at: str


class SupplierFunnelStepSummary(BaseModel):
    step: str
    viewed: int
    completed: int
    completion_rate_pct: float


class SupplierFunnelSummaryResponse(BaseModel):
    steps: list[SupplierFunnelStepSummary]
    total_step_views: int
    total_step_completions: int
    fda_exports_generated: int
    demo_resets_completed: int
    updated_at: str


class SupplierFDAExportRow(BaseModel):
    event_id: str
    tlc_code: str
    product_description: str | None = None
    cte_type: str
    facility_name: str
    event_time: str
    quantity: str
    unit_of_measure: str
    reference_document: str
    payload_sha256: str
    merkle_hash: str
    merkle_sequence: int


class SupplierFDAExportPreviewResponse(BaseModel):
    rows: list[SupplierFDAExportRow]
    total_count: int


FDA_EXPORT_COLUMN_ORDER = [
    "event_id",
    "tlc_code",
    "product_description",
    "cte_type",
    "facility_name",
    "event_time",
    "quantity",
    "unit_of_measure",
    "reference_document",
    "payload_sha256",
    "merkle_hash",
    "merkle_sequence",
]


FDA_EXPORT_HEADERS = {
    "event_id": "Event ID",
    "tlc_code": "Traceability Lot Code",
    "product_description": "Product Description",
    "cte_type": "Critical Tracking Event",
    "facility_name": "Facility",
    "event_time": "Event Time (UTC)",
    "quantity": "Quantity",
    "unit_of_measure": "Unit of Measure",
    "reference_document": "Reference Document",
    "payload_sha256": "SHA-256",
    "merkle_hash": "Merkle Hash",
    "merkle_sequence": "Merkle Sequence",
}


def _get_supplier_facility_or_404(
    db: Session,
    *,
    tenant_id: uuid_module.UUID,
    supplier_user_id: uuid_module.UUID,
    facility_id: uuid_module.UUID,
) -> SupplierFacilityModel:
    facility = db.get(SupplierFacilityModel, facility_id)
    if (
        facility is None
        or facility.tenant_id != tenant_id
        or facility.supplier_user_id != supplier_user_id
    ):
        raise HTTPException(status_code=404, detail="Facility not found")
    return facility


def _derive_required_ctes(categories: list[dict[str, Any]]) -> list[str]:
    required_ctes: list[str] = []
    seen: set[str] = set()
    for category in categories:
        for cte in category.get("ctes", []):
            if cte in seen:
                continue
            seen.add(cte)
            required_ctes.append(cte)
    return required_ctes


def _iso_utc(value: datetime) -> str:
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc).isoformat()
    return value.astimezone(timezone.utc).isoformat()


def _as_utc(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _merkle_integrity_ratio(events: list[SupplierCTEEventModel]) -> float:
    if not events:
        return 1.0

    checks = 0
    valid = 0
    previous: SupplierCTEEventModel | None = None

    for event in events:
        checks += 1
        if previous is None:
            if int(event.sequence_number) == 1 and event.merkle_prev_hash is None:
                valid += 1
        else:
            expected_sequence = int(previous.sequence_number) + 1
            if int(event.sequence_number) == expected_sequence and event.merkle_prev_hash == previous.merkle_hash:
                valid += 1
        previous = event

    return valid / checks if checks else 1.0


def _compute_supplier_compliance(
    db: Session,
    *,
    tenant_id: uuid_module.UUID,
    supplier_user_id: uuid_module.UUID,
    facility_id: uuid_module.UUID | None,
    lookback_days: int,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    facilities_query = select(SupplierFacilityModel).where(
        SupplierFacilityModel.tenant_id == tenant_id,
        SupplierFacilityModel.supplier_user_id == supplier_user_id,
    )
    if facility_id is not None:
        facilities_query = facilities_query.where(SupplierFacilityModel.id == facility_id)

    facilities = db.execute(facilities_query).scalars().all()
    evaluated_at = datetime.now(timezone.utc)
    evaluated_at_iso = _iso_utc(evaluated_at)
    if not facilities:
        return (
            {
                "score": 0,
                "coverage_ratio": 0.0,
                "freshness_ratio": 0.0,
                "integrity_ratio": 1.0,
                "required_ctes": 0,
                "covered_ctes": 0,
                "stale_ctes": 0,
                "missing_ctes": 0,
                "total_events": 0,
                "evaluated_at": evaluated_at_iso,
            },
            [],
        )

    facility_ids = [facility.id for facility in facilities]

    category_rows = db.execute(
        select(SupplierFacilityFTLCategoryModel).where(
            SupplierFacilityFTLCategoryModel.tenant_id == tenant_id,
            SupplierFacilityFTLCategoryModel.facility_id.in_(facility_ids),
        )
    ).scalars().all()

    scoped_events = db.execute(
        select(SupplierCTEEventModel)
        .where(
            SupplierCTEEventModel.tenant_id == tenant_id,
            SupplierCTEEventModel.supplier_user_id == supplier_user_id,
            SupplierCTEEventModel.facility_id.in_(facility_ids),
        )
        .order_by(SupplierCTEEventModel.sequence_number.asc())
    ).scalars().all()

    all_supplier_events = db.execute(
        select(SupplierCTEEventModel)
        .where(
            SupplierCTEEventModel.tenant_id == tenant_id,
            SupplierCTEEventModel.supplier_user_id == supplier_user_id,
        )
        .order_by(SupplierCTEEventModel.sequence_number.asc())
    ).scalars().all()

    required_by_facility: dict[uuid_module.UUID, list[str]] = {facility.id: [] for facility in facilities}
    required_seen_by_facility: dict[uuid_module.UUID, set[str]] = {facility.id: set() for facility in facilities}

    for row in category_rows:
        for cte in row.required_ctes or []:
            normalized_cte = str(cte).strip().lower()
            if not normalized_cte:
                continue
            if normalized_cte in required_seen_by_facility[row.facility_id]:
                continue
            required_seen_by_facility[row.facility_id].add(normalized_cte)
            required_by_facility[row.facility_id].append(normalized_cte)

    events_by_facility: dict[uuid_module.UUID, list[SupplierCTEEventModel]] = {facility.id: [] for facility in facilities}
    for event in scoped_events:
        events_by_facility[event.facility_id].append(event)

    threshold = evaluated_at - timedelta(days=lookback_days)
    required_total = 0
    covered_total = 0
    fresh_total = 0
    gap_payloads: list[dict[str, Any]] = []

    for facility in facilities:
        required_ctes = required_by_facility.get(facility.id, [])
        if not required_ctes:
            gap_payloads.append(
                {
                    "facility_id": str(facility.id),
                    "facility_name": facility.name,
                    "cte_type": None,
                    "severity": "high",
                    "issue": "No FTL categories scoped for this facility",
                    "reason": "ftl_not_scoped",
                    "last_seen": None,
                }
            )
            continue

        required_total += len(required_ctes)
        latest_by_cte: dict[str, datetime | None] = {}
        for event in events_by_facility.get(facility.id, []):
            normalized_cte = (event.cte_type or "").strip().lower()
            if not normalized_cte:
                continue
            event_time = _as_utc(event.event_time)
            existing = latest_by_cte.get(normalized_cte)
            if existing is None:
                latest_by_cte[normalized_cte] = event_time
                continue
            if event_time is not None and event_time > existing:
                latest_by_cte[normalized_cte] = event_time

        for cte in required_ctes:
            latest = latest_by_cte.get(cte)
            if latest is None:
                gap_payloads.append(
                    {
                        "facility_id": str(facility.id),
                        "facility_name": facility.name,
                        "cte_type": cte,
                        "severity": "high",
                        "issue": f"Missing required {cte} event",
                        "reason": "required_cte_missing",
                        "last_seen": None,
                    }
                )
                continue

            covered_total += 1
            if latest >= threshold:
                fresh_total += 1
            else:
                gap_payloads.append(
                    {
                        "facility_id": str(facility.id),
                        "facility_name": facility.name,
                        "cte_type": cte,
                        "severity": "medium",
                        "issue": f"{cte} event is older than {lookback_days} days",
                        "reason": "required_cte_stale",
                        "last_seen": _iso_utc(latest),
                    }
                )

    integrity_ratio = _merkle_integrity_ratio(all_supplier_events)

    if required_total == 0:
        coverage_ratio = 0.0
        freshness_ratio = 0.0
        score = 0
    else:
        coverage_ratio = covered_total / required_total
        freshness_ratio = fresh_total / required_total
        score = int(round(100 * ((coverage_ratio * 0.75) + (freshness_ratio * 0.15) + (integrity_ratio * 0.10))))
        score = max(0, min(100, score))

    score_payload = {
        "score": score,
        "coverage_ratio": round(coverage_ratio, 4),
        "freshness_ratio": round(freshness_ratio, 4),
        "integrity_ratio": round(integrity_ratio, 4),
        "required_ctes": required_total,
        "covered_ctes": covered_total,
        "stale_ctes": max(covered_total - fresh_total, 0),
        "missing_ctes": max(required_total - covered_total, 0),
        "total_events": len(scoped_events),
        "evaluated_at": evaluated_at_iso,
    }
    return score_payload, gap_payloads


def _compute_social_proof(db: Session, *, tenant_id: uuid_module.UUID) -> dict[str, Any]:
    suppliers_onboarded = int(
        db.execute(
            select(func.count(func.distinct(SupplierFacilityModel.supplier_user_id))).where(
                SupplierFacilityModel.tenant_id == tenant_id,
            )
        ).scalar_one()
        or 0
    )

    facilities_registered = int(
        db.execute(
            select(func.count(SupplierFacilityModel.id)).where(
                SupplierFacilityModel.tenant_id == tenant_id,
            )
        ).scalar_one()
        or 0
    )

    tlcs_tracked = int(
        db.execute(
            select(func.count(SupplierTraceabilityLotModel.id)).where(
                SupplierTraceabilityLotModel.tenant_id == tenant_id,
            )
        ).scalar_one()
        or 0
    )

    cte_events_verified = int(
        db.execute(
            select(func.count(SupplierCTEEventModel.id)).where(
                SupplierCTEEventModel.tenant_id == tenant_id,
            )
        ).scalar_one()
        or 0
    )

    fda_exports_generated = int(
        db.execute(
            select(func.count(SupplierFunnelEventModel.id)).where(
                SupplierFunnelEventModel.tenant_id == tenant_id,
                SupplierFunnelEventModel.event_name == "fda_export_downloaded",
            )
        ).scalar_one()
        or 0
    )

    return {
        "suppliers_onboarded": suppliers_onboarded,
        "facilities_registered": facilities_registered,
        "tlcs_tracked": tlcs_tracked,
        "cte_events_verified": cte_events_verified,
        "fda_exports_generated": fda_exports_generated,
        "updated_at": _iso_utc(datetime.now(timezone.utc)),
    }


SUPPLIER_FUNNEL_STEP_ORDER = [
    "buyer_invite",
    "supplier_signup",
    "facility_setup",
    "ftl_scoping",
    "cte_capture",
    "tlc_mgmt",
    "dashboard",
    "fda_export",
]


def _compute_funnel_summary(db: Session, *, tenant_id: uuid_module.UUID) -> dict[str, Any]:
    grouped_rows = db.execute(
        select(
            SupplierFunnelEventModel.step,
            SupplierFunnelEventModel.event_name,
            func.count(SupplierFunnelEventModel.id),
        )
        .where(SupplierFunnelEventModel.tenant_id == tenant_id)
        .group_by(SupplierFunnelEventModel.step, SupplierFunnelEventModel.event_name)
    ).all()

    viewed_by_step: dict[str, int] = {}
    completed_by_step: dict[str, int] = {}
    for step, event_name, count_value in grouped_rows:
        normalized_step = (step or "").strip().lower()
        if not normalized_step:
            continue
        normalized_event = (event_name or "").strip().lower()
        count_int = int(count_value or 0)
        if normalized_event == "step_viewed":
            viewed_by_step[normalized_step] = viewed_by_step.get(normalized_step, 0) + count_int
        elif normalized_event == "step_completed":
            completed_by_step[normalized_step] = completed_by_step.get(normalized_step, 0) + count_int

    ordered_steps = list(SUPPLIER_FUNNEL_STEP_ORDER)
    discovered_steps = sorted({*viewed_by_step.keys(), *completed_by_step.keys()} - set(ordered_steps))
    all_steps = ordered_steps + discovered_steps

    step_summaries: list[dict[str, Any]] = []
    total_views = 0
    total_completions = 0
    for step in all_steps:
        viewed = viewed_by_step.get(step, 0)
        completed = completed_by_step.get(step, 0)
        total_views += viewed
        total_completions += completed
        completion_rate_pct = round((completed / viewed) * 100, 1) if viewed > 0 else 0.0
        step_summaries.append(
            {
                "step": step,
                "viewed": viewed,
                "completed": completed,
                "completion_rate_pct": completion_rate_pct,
            }
        )

    fda_exports_generated = int(
        db.execute(
            select(func.count(SupplierFunnelEventModel.id)).where(
                SupplierFunnelEventModel.tenant_id == tenant_id,
                SupplierFunnelEventModel.event_name == "fda_export_downloaded",
            )
        ).scalar_one()
        or 0
    )
    demo_resets_completed = int(
        db.execute(
            select(func.count(SupplierFunnelEventModel.id)).where(
                SupplierFunnelEventModel.tenant_id == tenant_id,
                SupplierFunnelEventModel.event_name == "demo_reset_completed",
            )
        ).scalar_one()
        or 0
    )

    return {
        "steps": step_summaries,
        "total_step_views": total_views,
        "total_step_completions": total_completions,
        "fda_exports_generated": fda_exports_generated,
        "demo_resets_completed": demo_resets_completed,
        "updated_at": _iso_utc(datetime.now(timezone.utc)),
    }


def _string_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def _build_fda_export_rows(
    db: Session,
    *,
    tenant_id: uuid_module.UUID,
    supplier_user_id: uuid_module.UUID,
    facility_id: uuid_module.UUID | None,
    tlc_code: str | None,
    start_time: datetime | None,
    end_time: datetime | None,
) -> list[dict[str, Any]]:
    event_query = (
        select(
            SupplierCTEEventModel,
            SupplierTraceabilityLotModel,
            SupplierFacilityModel,
        )
        .join(SupplierTraceabilityLotModel, SupplierTraceabilityLotModel.id == SupplierCTEEventModel.lot_id)
        .join(SupplierFacilityModel, SupplierFacilityModel.id == SupplierCTEEventModel.facility_id)
        .where(
            SupplierCTEEventModel.tenant_id == tenant_id,
            SupplierCTEEventModel.supplier_user_id == supplier_user_id,
            SupplierTraceabilityLotModel.tenant_id == tenant_id,
            SupplierTraceabilityLotModel.supplier_user_id == supplier_user_id,
            SupplierFacilityModel.tenant_id == tenant_id,
            SupplierFacilityModel.supplier_user_id == supplier_user_id,
        )
    )

    if facility_id is not None:
        event_query = event_query.where(SupplierCTEEventModel.facility_id == facility_id)
    if tlc_code:
        event_query = event_query.where(SupplierTraceabilityLotModel.tlc_code == tlc_code)

    normalized_start_time = _as_utc(start_time)
    normalized_end_time = _as_utc(end_time)
    if normalized_start_time is not None:
        event_query = event_query.where(SupplierCTEEventModel.event_time >= normalized_start_time)
    if normalized_end_time is not None:
        event_query = event_query.where(SupplierCTEEventModel.event_time <= normalized_end_time)

    event_rows = db.execute(
        event_query.order_by(SupplierCTEEventModel.event_time.desc(), SupplierCTEEventModel.sequence_number.desc())
    ).all()

    output_rows: list[dict[str, Any]] = []
    for event, lot, facility in event_rows:
        kde_data = event.kde_data if isinstance(event.kde_data, dict) else {}
        quantity_value = kde_data.get("quantity", kde_data.get("qty", ""))
        unit_value = kde_data.get("unit_of_measure", kde_data.get("uom", ""))
        reference_document = (
            kde_data.get("reference_document")
            or kde_data.get("reference_document_id")
            or kde_data.get("bill_of_lading")
            or kde_data.get("reference_doc")
            or ""
        )

        output_rows.append(
            {
                "event_id": str(event.id),
                "tlc_code": lot.tlc_code,
                "product_description": lot.product_description or kde_data.get("product_description"),
                "cte_type": event.cte_type,
                "facility_name": facility.name,
                "event_time": _iso_utc(_as_utc(event.event_time) or datetime.now(timezone.utc)),
                "quantity": _string_value(quantity_value),
                "unit_of_measure": _string_value(unit_value),
                "reference_document": _string_value(reference_document),
                "payload_sha256": event.payload_sha256,
                "merkle_hash": event.merkle_hash,
                "merkle_sequence": int(event.sequence_number),
            }
        )

    return output_rows


def _render_fda_export_csv(rows: list[dict[str, Any]]) -> bytes:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[FDA_EXPORT_HEADERS[column] for column in FDA_EXPORT_COLUMN_ORDER])
    writer.writeheader()

    for row in rows:
        writer.writerow(
            {
                FDA_EXPORT_HEADERS[column]: _string_value(row.get(column, ""))
                for column in FDA_EXPORT_COLUMN_ORDER
            }
        )

    return output.getvalue().encode("utf-8")


def _xlsx_column_name(index: int) -> str:
    name = ""
    current = index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _xml_escape(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def _render_fda_export_xlsx_fallback(rows: list[dict[str, Any]]) -> bytes:
    header_values = [FDA_EXPORT_HEADERS[column] for column in FDA_EXPORT_COLUMN_ORDER]
    last_column = _xlsx_column_name(len(FDA_EXPORT_COLUMN_ORDER))

    sheet_rows: list[str] = []

    header_cells = []
    for col_idx, header in enumerate(header_values, start=1):
        cell_ref = f"{_xlsx_column_name(col_idx)}1"
        header_cells.append(
            f"<c r=\"{cell_ref}\" t=\"inlineStr\"><is><t>{_xml_escape(header)}</t></is></c>"
        )
    sheet_rows.append(f"<row r=\"1\">{''.join(header_cells)}</row>")

    for row_idx, row in enumerate(rows, start=2):
        data_cells: list[str] = []
        for col_idx, column in enumerate(FDA_EXPORT_COLUMN_ORDER, start=1):
            cell_ref = f"{_xlsx_column_name(col_idx)}{row_idx}"
            value = _xml_escape(_string_value(row.get(column, "")))
            data_cells.append(
                f"<c r=\"{cell_ref}\" t=\"inlineStr\"><is><t>{value}</t></is></c>"
            )
        sheet_rows.append(f"<row r=\"{row_idx}\">{''.join(data_cells)}</row>")

    cols_xml_parts: list[str] = []
    for idx, column in enumerate(FDA_EXPORT_COLUMN_ORDER, start=1):
        display_name = FDA_EXPORT_HEADERS[column]
        width = min(max(len(display_name) + 2, 14), 42)
        cols_xml_parts.append(f"<col min=\"{idx}\" max=\"{idx}\" width=\"{width}\" customWidth=\"1\"/>")

    worksheet_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<worksheet xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\">"
        "<sheetViews><sheetView workbookViewId=\"0\"><pane ySplit=\"1\" topLeftCell=\"A2\" activePane=\"bottomLeft\" state=\"frozen\"/>"
        "</sheetView></sheetViews>"
        "<sheetFormatPr defaultRowHeight=\"15\"/>"
        f"<cols>{''.join(cols_xml_parts)}</cols>"
        f"<sheetData>{''.join(sheet_rows)}</sheetData>"
        f"<autoFilter ref=\"A1:{last_column}1\"/>"
        "</worksheet>"
    )

    workbook_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<workbook xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\" "
        "xmlns:r=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships\">"
        "<sheets><sheet name=\"FDA Traceability\" sheetId=\"1\" r:id=\"rId1\"/></sheets>"
        "</workbook>"
    )

    root_rels_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">"
        "<Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument\" Target=\"xl/workbook.xml\"/>"
        "</Relationships>"
    )

    workbook_rels_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">"
        "<Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet\" Target=\"worksheets/sheet1.xml\"/>"
        "</Relationships>"
    )

    content_types_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Types xmlns=\"http://schemas.openxmlformats.org/package/2006/content-types\">"
        "<Default Extension=\"rels\" ContentType=\"application/vnd.openxmlformats-package.relationships+xml\"/>"
        "<Default Extension=\"xml\" ContentType=\"application/xml\"/>"
        "<Override PartName=\"/xl/workbook.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml\"/>"
        "<Override PartName=\"/xl/worksheets/sheet1.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml\"/>"
        "</Types>"
    )

    payload = io.BytesIO()
    with zipfile.ZipFile(payload, mode="w", compression=zipfile.ZIP_DEFLATED) as workbook_zip:
        workbook_zip.writestr("[Content_Types].xml", content_types_xml)
        workbook_zip.writestr("_rels/.rels", root_rels_xml)
        workbook_zip.writestr("xl/workbook.xml", workbook_xml)
        workbook_zip.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        workbook_zip.writestr("xl/worksheets/sheet1.xml", worksheet_xml)
    return payload.getvalue()


def _render_fda_export_xlsx(rows: list[dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        return _render_fda_export_xlsx_fallback(rows)

    workbook = Workbook()
    worksheet: Any = workbook.active
    worksheet.title = "FDA Traceability"
    worksheet.append([FDA_EXPORT_HEADERS[column] for column in FDA_EXPORT_COLUMN_ORDER])

    for row in rows:
        worksheet.append([_string_value(row.get(column, "")) for column in FDA_EXPORT_COLUMN_ORDER])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for index, column_name in enumerate(FDA_EXPORT_COLUMN_ORDER, start=1):
        display_name = FDA_EXPORT_HEADERS[column_name]
        width = max(len(display_name) + 2, 14)
        worksheet.column_dimensions[get_column_letter(index)].width = min(width, 42)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Route handlers have been decomposed into focused sub-modules:
#   - supplier_facilities_routes.py  (Facilities CRUD, FTL categories, CTEs, TLCs)
#   - supplier_compliance_routes.py  (Compliance score, gaps, FDA export)
#   - supplier_funnel_routes.py      (Funnel events, social proof, demo reset)
#
# This file retains all shared models, Pydantic schemas, helper functions,
# and constants that the sub-routers import. The `router` object is kept
# (empty) so that existing imports of `from app.supplier_onboarding_routes
# import router` continue to work without breaking downstream code.
# ---------------------------------------------------------------------------
